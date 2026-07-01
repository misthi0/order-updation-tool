from flask import Flask, request, jsonify
from flask_cors import CORS
import msoffcrypto
import openpyxl
import io
import os
from datetime import datetime, timedelta

app = Flask(__name__)
CORS(app)

GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')

def safe_str(val):
    return str(val).strip() if val is not None else ''

def to_date_str(value):
    if isinstance(value, datetime):
        return value.strftime('%d-%m-%Y')
    if isinstance(value, (int, float)) and 40000 < value < 60000:
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(value))).strftime('%d-%m-%Y')
        except:
            pass
    s = safe_str(value)
    if not s or s == 'None':
        return ''
    return s.replace('.', '-').replace('/', '-')

KNOWN_PLACES = [
    'kalamassery', 'kalamassary', 'perambara', 'perambra',
    'spinmax', 'spinnmax', 'chennai', 'baroda', 'limda', 'linda',
    'pune', 'ace', 'ap'
]

# Map from header name → SAP plant code
# These are the column headers at the far right of the Excel
PLANT_HEADER_MAP = {
    'gmpd':    'GMPD',
    'ptg':     'PTG',
    'rnkt':    'RNKT',
    'rkt':     'RKT',
    'ptg scm': 'PTG SCM',
    'inc1':    'INC1',
    'ing1':    'ING1',
    'inp1':    'INP1',
    'inr1':    'INR1',
}

def extract_location_from_sheet_name(sheet_name):
    lower = sheet_name.lower()
    for place in KNOWN_PLACES:
        if place in lower:
            return place
    part = sheet_name.split('-')[0].split('–')[0].strip()
    return part.lower() if part else ''

def extract_rows_directly(ws, location_override=''):
    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row containing 'grade' or 'material'
    header_idx = -1
    for i, row in enumerate(all_rows[:10]):
        row_lower = [safe_str(c).lower() for c in row]
        if any(c in ('grade', 'material', 'material number') for c in row_lower):
            header_idx = i
            break

    if header_idx == -1:
        return None, None

    headers = [safe_str(c).lower().strip() for c in all_rows[header_idx]]

    # Find indices of plant-code columns (gmpd, ptg, rnkt, ptg scm etc.)
    # These columns have a non-empty value when that plant supplies the row
    plant_header_indices = []
    for idx, h in enumerate(headers):
        if h in PLANT_HEADER_MAP:
            plant_header_indices.append((idx, PLANT_HEADER_MAP[h]))

    def col(row, *keys):
        for key in keys:
            for idx, h in enumerate(headers):
                if key in h and idx < len(row):
                    val = row[idx]
                    if val is not None and safe_str(val) not in ('', 'None'):
                        return val
        return ''

    def get_plant(row):
        # Step 1: Try "Supplying Plant" column (works for PTG/GMPD/RKT rows)
        sp = safe_str(col(row, 'supplying plant')).upper()
        if sp and sp not in ('', 'NONE'):
            return sp

        # Step 2: Scan plant-code columns (gmpd/ptg/rnkt/ptg scm at far right)
        # The column header IS the plant code; non-empty cell = that plant supplies this row
        for idx, plant_code in plant_header_indices:
            if idx < len(row):
                cell = row[idx]
                if cell is not None and safe_str(cell) not in ('', 'None', '0'):
                    return plant_code  # header name = plant code
                print(f"NO PLANT FOUND for row: grade={safe_str(col(row,'grade','material'))}, plant_header_indices={plant_header_indices}, row_values_at_plant_cols={[(idx, row[idx] if idx < len(row) else 'OUT_OF_RANGE') for idx, _ in plant_header_indices]}")
   

        return ''

    data_rows = all_rows[header_idx + 1:]

    # Location from sheet name
    if location_override:
        location = location_override
    else:
        loc_counts = {}
        for row in data_rows:
            l = safe_str(col(row, 'delivery location', 'location', 'destination')).lower()
            if l and l != 'none':
                loc_counts[l] = loc_counts.get(l, 0) + 1
        location = sorted(loc_counts, key=loc_counts.get, reverse=True)[0] if loc_counts else ''

    plant_counts = {}
    items = []

    for row in data_rows:
        grade     = safe_str(col(row, 'grade', 'material number', 'material'))
        pack      = safe_str(col(row, 'pack', 'packing', 'pack type'))
        qty_raw   = safe_str(col(row, 'schedule qty', 'quantity', 'qty', 'schedule quantity'))
        deliv_raw = col(row, 'delivery schedule', 'delivery date', 'dispatch schedule')
        amount    = safe_str(col(row, 'amount', 'price', 'basic price', 'rate'))
        row_plant = get_plant(row)

        if not grade or grade == 'None':
            continue

        deliv_date = to_date_str(deliv_raw) if deliv_raw else ''
        if not deliv_date:
            continue

        try:
            if float(qty_raw) <= 0:
                continue
        except:
            if not qty_raw:
                continue

        if row_plant:
            plant_counts[row_plant] = plant_counts.get(row_plant, 0) + 1

        items.append({
            'MATERIAL_NUMBER':  f'{grade} {pack}'.strip() if pack else grade,
            'QUANTITY':         qty_raw,
            'SALES_UNIT':       'MT',
            'DELIVERY_DATE':    deliv_date,
            'AMOUNT_1':         amount,
            'CONDITION_TYPE_1': '',
            'CONDITION_TYPE_2': '',
            'AMOUNT_2':         '',
            'SOLD_TO_PARTY':    location,
            'SHIP_TO_PARTY':    location,
            'PLANT':            row_plant,
        })

    most_common_plant = sorted(plant_counts, key=plant_counts.get, reverse=True)[0] if plant_counts else ''

    header = {
        'ORDER_TYPE':            'ZDOM',
        'SALES_ORG':             most_common_plant,
        'SOLD_TO_PARTY':         location,
        'SHIP_TO_PARTY':         location,
        'PO_NO':                 '',
        'PURCHASE_ORDER_DATE':   '',
        'REQ_DELIVERY_DATE':     '',
        'INCOTERM':              '',
        'INCOTERM2':             '',
        'ALT_TAX_CLASSF':        '',
        'CUSTOMER_GROUP':        '',
        'PAYER':                 '',
        'SPECIAL_STOCK_PARTNER': '',
        'PLANT':                 most_common_plant,
    }

    return header, items


def extract_with_groq(ws, api_key):
    import csv, requests, json

    rows = []
    for row in ws.iter_rows(values_only=True):
        converted = []
        for cell in row:
            if isinstance(cell, datetime):
                converted.append(cell.strftime('%d-%m-%Y'))
            elif isinstance(cell, (int, float)) and 40000 < cell < 60000:
                try:
                    converted.append((datetime(1899, 12, 30) + timedelta(days=int(cell))).strftime('%d-%m-%Y'))
                except:
                    converted.append(str(cell) if cell is not None else '')
            else:
                converted.append(str(cell) if cell is not None else '')
        if any(c.strip() for c in converted):
            rows.append(converted)

    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    csv_text = output.getvalue()

    prompt = f"""Extract ALL order line items. Return ONLY valid JSON, no markdown.

{{
  "ORDER_TYPE": "ZDOM or ZEXP",
  "SALES_ORG": "plant code",
  "SOLD_TO_PARTY": "customer name or code",
  "SHIP_TO_PARTY": "delivery location name",
  "PO_NO": "",
  "PURCHASE_ORDER_DATE": "DD-MM-YYYY",
  "REQ_DELIVERY_DATE": "DD-MM-YYYY",
  "INCOTERM": "",
  "INCOTERM2": "",
  "PLANT": "plant code",
  "MATERIAL_NUMBER": ["Grade Pack per row"],
  "QUANTITY": ["qty per row"],
  "SALES_UNIT": ["MT or KG per row"],
  "DELIVERY_DATE": ["DD-MM-YYYY per row"],
  "AMOUNT_1": ["price per row or empty"]
}}

CRITICAL: Extract EVERY row. Skip totals/headers/blanks. All arrays same length.

CSV:
{csv_text[:10000]}"""

    response = requests.post(
        'https://api.groq.com/openai/v1/chat/completions',
        headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}'},
        json={'model': 'llama-3.3-70b-versatile', 'max_tokens': 4000,
              'messages': [{'role': 'user', 'content': prompt}]}
    )

    data = response.json()
    if 'choices' not in data:
        raise Exception('Groq API error: ' + str(data))

    raw = data['choices'][0]['message']['content'].strip()
    cleaned = raw.replace('```json', '').replace('```', '').strip()

    parsed = json.loads(cleaned)

    # Guard: Groq sometimes returns a list instead of dict
    if isinstance(parsed, list):
        raise Exception('Groq returned a list instead of object')

    def to_list(v):
        if isinstance(v, list): return [str(x) for x in v]
        return [str(v)] if v else []

    header = {
        'ORDER_TYPE':            parsed.get('ORDER_TYPE', ''),
        'SALES_ORG':             parsed.get('SALES_ORG', ''),
        'SOLD_TO_PARTY':         parsed.get('SOLD_TO_PARTY', ''),
        'SHIP_TO_PARTY':         parsed.get('SHIP_TO_PARTY', ''),
        'PO_NO':                 parsed.get('PO_NO', ''),
        'PURCHASE_ORDER_DATE':   parsed.get('PURCHASE_ORDER_DATE', ''),
        'REQ_DELIVERY_DATE':     parsed.get('REQ_DELIVERY_DATE', ''),
        'INCOTERM':              parsed.get('INCOTERM', ''),
        'INCOTERM2':             '',
        'ALT_TAX_CLASSF':        '',
        'CUSTOMER_GROUP':        '',
        'PAYER':                 '',
        'SPECIAL_STOCK_PARTNER': '',
        'PLANT':                 parsed.get('PLANT', ''),
    }

    materials  = to_list(parsed.get('MATERIAL_NUMBER', []))
    quantities = to_list(parsed.get('QUANTITY', []))
    units      = to_list(parsed.get('SALES_UNIT', []))
    dates      = to_list(parsed.get('DELIVERY_DATE', []))
    amounts    = to_list(parsed.get('AMOUNT_1', []))

    items = []
    for i, mat in enumerate(materials):
        if not mat.strip(): continue
        items.append({
            'MATERIAL_NUMBER':  mat,
            'QUANTITY':         quantities[i] if i < len(quantities) else '',
            'SALES_UNIT':       units[i] if i < len(units) else 'MT',
            'DELIVERY_DATE':    dates[i] if i < len(dates) else '',
            'AMOUNT_1':         amounts[i] if i < len(amounts) else '',
            'CONDITION_TYPE_1': '',
            'CONDITION_TYPE_2': '',
            'AMOUNT_2':         '',
        })

    return header, items


@app.route('/decrypt', methods=['POST'])
def decrypt_file():
    try:
        file = request.files['file']
        password = request.form.get('password', '')
        api_key = request.form.get('api_key', GROQ_API_KEY)
        file_data = file.read()

        try:
            office_file = msoffcrypto.OfficeFile(io.BytesIO(file_data))
            if office_file.is_encrypted():
                if not password:
                    return jsonify({'error': 'PASSWORD_REQUIRED'}), 400
                decrypted = io.BytesIO()
                office_file.load_key(password=password)
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                file_data = decrypted.read()
        except Exception as e:
            if 'password' in str(e).lower() or 'key' in str(e).lower():
                return jsonify({'error': 'WRONG_PASSWORD'}), 401

        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)

        all_items = []
        combined_header = None

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not ws.max_row or ws.max_row < 2:
                continue

            # Skip Summary sheet — it has no per-row order data
            if 'summary' in sheet_name.lower():
                continue

            location_hint = extract_location_from_sheet_name(sheet_name)
            header, items = extract_rows_directly(ws, location_override=location_hint)

            if header is None or not items:
                continue

            if combined_header is None:
                combined_header = header

            all_items.extend(items)

        if not combined_header:
            return jsonify({'error': 'No data found in any sheet'}), 400

        return jsonify({'success': True, 'header': combined_header, 'items': all_items})

    except Exception as e:
        error_msg = str(e)
        if 'password' in error_msg.lower() or 'decrypt' in error_msg.lower():
            return jsonify({'error': 'WRONG_PASSWORD'}), 401
        return jsonify({'error': error_msg}), 500


if __name__ == '__main__':
    app.run(port=5001, debug=True)